from __future__ import annotations
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

try:
    from openpyxl.drawing.image import Image as ExcelImage
except ImportError:
    ExcelImage = None

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
except ImportError:
    CellRichText = None
    TextBlock = None
    InlineFont = None

from app.config import Settings
from app.data_access import RequestFilters, STATUS_ORDER, _normalize_status_counts

EXCEL_POINTS_PER_PIXEL = 0.75
REQUEST_ROW_HEIGHT_PIXELS = 27
REQUEST_ROW_HEIGHT_POINTS = REQUEST_ROW_HEIGHT_PIXELS * EXCEL_POINTS_PER_PIXEL
ERROR_COMMENT_COLUMN = "Bot Error Comment"
ERROR_COMMENT_COLUMN_WIDTH = 70
ERROR_COMMENT_WRAP_CHARS = 70
ERROR_COMMENT_MAX_ROW_HEIGHT = 150
ERROR_COMMENT_LABEL_COLOR = "C0504D"
ERROR_COMMENT_BASE_COLOR = "244062"

LOG_FINAL_STATUS_KEY = "LogFinalAppStatus"
LOG_MISSING_PROSPECT_FIELDS_KEY = "LogMissingProspectFields"
LOG_MISSING_PROSPECT_DOCS_KEY = "LogMissingProspectDocs"
LOG_MISSING_LEAD_FIELDS_KEY = "LogMissingLeadFields"
LOG_MISSING_LEAD_DOCS_KEY = "LogMissingLeadDocs"

BRAND_COLORS = {
    "brand_blue": "00385F",
    "slate": "1E293B",
    "white": "FFFFFF",
    "panel": "FFFFFF",
    "border": "E2E8F0",
    "text": "0F172A",
    "muted": "64748B",
    "success": "15803D",
    "failed": "B91C1C",
    "success_bg": "DCFCE7",
    "failed_bg": "FEE2E2",
}

STATUS_STYLES = {
    "Success": (BRAND_COLORS["success_bg"], BRAND_COLORS["success"]),
    "Failed": (BRAND_COLORS["failed_bg"], BRAND_COLORS["failed"]),
}

EMAIL_REQUEST_FIELDS = [
    "Outlet Name",
    "Request ID",
    "Emirate",
    "Email",
    "Mobile Number",
    "Vehicle CV",
    "Registration Emirates",
    "Bank Financed Y/N",
    "Bank Name",
    "No of Documents",
    "Submitted ON",
    "Bot Status(Passed/Error)",
    "Bot Error Comment",
    "Pass Status(Lead / Prospect)",
    "Lead_Ref_No",
    "Prospect_Ref_No",
    "CRM_Client Type *",
    "CRM_First Name *",
    "CRM_Last Name *",
    "CRM_Mobile No. *",
    "CRM_Email Id",
    "CRM_Date of Birth",
    "CRM_Business Type *",
    "CRM_Class *",
    "CRM_Policy Type",
    "CRM_License Issue Date",
    "CRM_License Number",
    "CRM_License Expiry Date",
    "CRM_License Issue Place",
    "CRM_Make *",
    "CRM_Model *",
    "CRM_Year Of Manufacture *",
    "CRM_Vehicle Colour",
    "CRM_Engine Numberc",
    "CRM_Chassis Number",
    "CRM_Vehicle Value",
    "CRM_Regn No.",
    "CRM_Date of First Registration",
    "CRM_Emirate",
    "CRM_TCF No",
]

EMAIL_FIELD_ALIASES = {
    "Outlet Name": ["OutletName", "POS", "Location"],
    "Request ID": ["RequestId", "RequestID", "request_id"],
    "Emirate": ["Emirate"],
    "Email": ["Email_CRM"],
    "Mobile Number": ["Mobile_CRM"],
    "Vehicle CV": ["VehicleValue"],
    "Registration Emirates": ["Emirate"],
    "Bank Financed Y/N": ["BankFinancedYN", "BankFinanced"],
    "Bank Name": ["BankName"],
    "No of Documents": ["NoOfDocuments", "DocumentCount", "DocumentsCount"],
    "Submitted ON": ["SubmittedOn", "SubmittedDate", "CreatedAt", "GenerationDate"],
    "Bot Status(Passed/Error)": ["BotStatus", "ValidationStatus", "CRMStatus", "AppStatus"],
    "Bot Error Comment": ["ValidationError","LastError"],
    "Pass Status(Lead / Prospect)": ["AppStatus"],
    "Lead_Ref_No": ["LeadRefNo"],
    "Prospect_Ref_No": ["ProspectRefNo"],
    "CRM_Client Type *": ["ClientType"],
    "CRM_First Name *": ["FirstName"],
    "CRM_Last Name *": ["LastName"],
    "CRM_Mobile No. *": ["Mobile_CRM", "InsuredMobileNumber", "PhoneNumber"],
    "CRM_Email Id": ["Email_CRM", "Email"],
    "CRM_Date of Birth": ["DateOfBirth", "DOB"],
    "CRM_Business Type *": ["BusinessType"],
    "CRM_Class *": ["Class", "VehicleClass"],
    "CRM_Policy Type": ["PolicyTypeCRM", "PolicyType"],
    "CRM_License Issue Date": ["LicenseIssueDate"],
    "CRM_License Number": ["LicenseNumber", "LicenseNo"],
    "CRM_License Expiry Date": ["LicenseExpiryDate"],
    "CRM_License Issue Place": ["LicenseIssuePlace"],
    "CRM_Make *": ["Make", "VehicleMake"],
    "CRM_Model *": ["Model", "VehicleModel"],
    "CRM_Year Of Manufacture *": ["YearOfManufacture", "ManufactureYear"],
    "CRM_Vehicle Colour": ["VehicleColor", "VehicleColour", "Colour", "Color"],
    "CRM_Engine Numberc": ["EngineNumber", "EngineNo"],
    "CRM_Chassis Number": ["ChassisNumber", "ChassisNo"],
    "CRM_Vehicle Value": ["VehicleValue", "Amount"],
    "CRM_Regn No.": ["VehiclePlateNumber", "RegistrationNo", "RegistrationNumber", "RegnNo"],
    "CRM_Date of First Registration": ["DateOfFirstRegistration", "FirstRegistrationDate"],
    "CRM_Emirate": ["Emirate"],
    "CRM_TCF No": ["TCFNo", "TcfNo"],
}

EMAIL_COMPUTED_FIELDS = {"Bank Financed Y/N", "Vehicle CV", "Registration Emirates"}

EMAIL_DATE_FIELDS = {
    "Submitted ON",
    "CRM_Date of Birth",
    "CRM_License Issue Date",
    "CRM_License Expiry Date",
    "CRM_Date of First Registration",
}

class MissingEmailFieldsError(ValueError):
    def __init__(self, missing_fields: list[str], available_columns: list[str]):
        self.missing_fields = missing_fields
        self.available_columns = available_columns
        super().__init__(
            "Missing email template fields: "
            + ", ".join(missing_fields)
            + ". Available columns: "
            + ", ".join(available_columns)
        )

def build_excel(
    rows: list[dict[str, Any]],
    filters: RequestFilters,
    settings: Settings,
    export_profile: str = "full",
    allow_missing_email_fields: bool = False,
) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    counts = _status_counts(rows, settings)
    _build_summary_sheet(summary, rows, filters, counts, settings)

    requests = wb.create_sheet(_safe_sheet_name("Requests"))
    if export_profile == "email":
        headers, request_rows = _email_template_table(rows, allow_missing=allow_missing_email_fields)
        _build_requests_sheet(requests, request_rows, settings, columns=headers, uppercase_headers=False)
    else:
        _build_requests_sheet(requests, rows, settings)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()

def _build_summary_sheet(ws, rows: list[dict[str, Any]], filters: RequestFilters, counts: dict[str, int], settings: Settings) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    
    _style_cells(ws, "A1:I40", fill=BRAND_COLORS["white"])

    # Columns fixed to 9 total
    for column, width in {"A": 22, "B": 15, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15, "H": 15, "I": 15}.items():
        ws.column_dimensions[column].width = width

    for row_num in range(1, 32):
        ws.row_dimensions[row_num].height = 24

    _build_header(ws, settings)
    _build_kpi_cards(ws, len(rows), counts)
    _build_footer(ws)

def _build_header(ws, settings: Settings) -> None:
    _style_cells(ws, "A1:I3", fill=BRAND_COLORS["white"], border=BRAND_COLORS["border"])

    # Logo in Column A only (Row 1 to 3)
    ws.merge_cells("A1:A3")

    # Title in Column B to I
    ws.merge_cells("B1:I3")
    ws["B1"] = "CRM Monitor Report"
    ws["B1"].font = Font(name="Inter", size=22, bold=True, color=BRAND_COLORS["brand_blue"])
    ws["B1"].alignment = Alignment(vertical="center", horizontal="left")

    _add_logo(ws, settings, "A1")

def _build_kpi_cards(ws, total: int, counts: dict[str, int]) -> None:
    cards = [
        ("A5:C8", "Total Requests", total, BRAND_COLORS["slate"]),
        ("D5:F8", "Success", counts.get("Success", 0), BRAND_COLORS["success"]),
        ("G5:I8", "Failed", counts.get("Failed", 0), BRAND_COLORS["failed"]),
    ]

    for cell_range, label, value, text_color in cards:
        start_cell = cell_range.split(":")[0]
        col = "".join(filter(str.isalpha, start_cell))
        row = int("".join(filter(str.isdigit, start_cell)))

        _style_cells(ws, cell_range, fill=BRAND_COLORS["panel"], border=BRAND_COLORS["border"])
        ws.merge_cells(start_row=row, start_column=ws[col + str(row)].column, end_row=row + 1, end_column=ws[col + str(row)].column + 2)
        ws.merge_cells(start_row=row + 2, start_column=ws[col + str(row)].column, end_row=row + 3, end_column=ws[col + str(row)].column + 2)

        ws[f"{col}{row}"] = label.upper()
        ws[f"{col}{row}"].font = Font(name="Inter", size=10, bold=True, color=BRAND_COLORS["muted"])
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws[f"{col}{row + 2}"] = value
        ws[f"{col}{row + 2}"].font = Font(name="Inter", size=26, bold=True, color=text_color)
        ws[f"{col}{row + 2}"].alignment = Alignment(horizontal="center", vertical="center")

def _build_footer(ws) -> None:
    ws.merge_cells("A10:I10")
    ws["A10"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A10"].font = Font(name="Inter", size=10, color=BRAND_COLORS["muted"])
    ws["A10"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A11:I11")
    ws["A11"] = "Prepared By Data Processing Team | Algospring"
    ws["A11"].font = Font(name="Inter", size=10, color=BRAND_COLORS["muted"], italic=True)
    ws["A11"].alignment = Alignment(horizontal="center")

def _build_requests_sheet(
    ws,
    rows: list[dict[str, Any]] | list[list[Any]],
    settings: Settings,
    columns: list[str] | None = None,
    uppercase_headers: bool = True,
) -> None:
    columns = columns or (list(rows[0].keys()) if rows else ["No data matched the selected filters"])
    ws.append([str(c).upper() if uppercase_headers else str(c) for c in columns])
    for row in rows:
        if isinstance(row, dict):
            ws.append([row.get(column) for column in columns])
        else:
            ws.append(row)
    status_col_idx = _status_column_index(columns, settings)
    _apply_requests_table_style(ws, max(1, ws.max_row), max(1, ws.max_column), status_col_idx)

def _apply_requests_table_style(ws, max_row: int, max_col: int, status_col_idx: int | None) -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    thin = Side(style="thin", color=BRAND_COLORS["border"])
    error_comment_col_idx = _header_column_index(ws, ERROR_COMMENT_COLUMN)
    for row_idx in range(1, max_row + 1):
        ws.row_dimensions[row_idx].height = REQUEST_ROW_HEIGHT_POINTS

    for cell in ws[1]:
        cell.fill = _fill(BRAND_COLORS["slate"])
        cell.font = Font(name="Inter", size=10, color=BRAND_COLORS["white"], bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    row_fill_white = _fill(BRAND_COLORS["white"])
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.fill = row_fill_white
            cell.border = Border(bottom=thin, top=thin, left=thin, right=thin)
            if error_comment_col_idx and cell.column == error_comment_col_idx:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="center")
            cell.font = Font(name="Inter", color=BRAND_COLORS["text"])
        if error_comment_col_idx:
            comment_cell = row[error_comment_col_idx - 1]
            wrapped_lines = _wrapped_line_count(comment_cell.value)
            ws.row_dimensions[comment_cell.row].height = min(
                ERROR_COMMENT_MAX_ROW_HEIGHT,
                max(REQUEST_ROW_HEIGHT_POINTS, wrapped_lines * 15),
            )
        if status_col_idx:
            status_cell = row[status_col_idx - 1]
            fill, font_color = STATUS_STYLES.get(_canonical_status(status_cell.value), (None, None))
            if fill and font_color:
                status_cell.fill = _fill(fill)
                status_cell.font = Font(name="Inter", bold=True, color=font_color)
                status_cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        if error_comment_col_idx and col_idx == error_comment_col_idx:
            ws.column_dimensions[letter].width = ERROR_COMMENT_COLUMN_WIDTH
            continue
        max_length = 12
        for cell in ws[letter]:
            max_length = max(max_length, min(len(_plain_text(cell.value)), 45))
        ws.column_dimensions[letter].width = max_length + 3

def _header_column_index(ws, target: str) -> int | None:
    normalized_target = _normalize_column(target)
    for index, cell in enumerate(ws[1], start=1):
        if _normalize_column(str(cell.value or "")) == normalized_target:
            return index
    return None

def _wrapped_line_count(value: Any) -> int:
    text = _plain_text(value)
    if not text:
        return 1
    return sum(max(1, (len(part) // ERROR_COMMENT_WRAP_CHARS) + 1) for part in text.splitlines() or [""])

def _status_counts(rows: list[dict[str, Any]], settings: Settings) -> dict[str, int]:
    counts: dict[str, int] = {}
    for row in rows:
        status = _status_value(row, settings)
        counts[status] = counts.get(status, 0) + 1
    return _normalize_status_counts(counts)

def _status_value(row: dict[str, Any], settings: Settings) -> str:
    for key in [
        settings.status_column,
        "ValidationStatus",
        "validationStatus",
        "validation_status",
        "CRMStatus",
        "crmStatus",
        "crm_status",
        "Status",
        "status",
    ]:
        if key in row:
            return _canonical_status(row.get(key))
    return "Unknown"

def _canonical_status(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return "Unknown"
    normalized = text.casefold().replace(" ", "").replace("_", "").replace("-", "")
    mapping = {"success": "Success", "failed": "Failed", "fail": "Failed", "error": "Failed", "pending": "Pending", "inprogress": "In Progress"}
    return mapping.get(normalized, text)

def _status_column_index(columns: list[str], settings: Settings) -> int | None:
    candidates = {
        settings.status_column,
        "ValidationStatus",
        "validationStatus",
        "validation_status",
        "CRMStatus",
        "crmStatus",
        "crm_status",
        "Status",
        "status",
    }
    for index, column in enumerate(columns, start=1):
        if column in candidates:
            return index
    return None

def _email_template_table(
    rows: list[dict[str, Any]],
    allow_missing: bool = False,
) -> tuple[list[str], list[list[Any]]]:
    if not rows:
        return EMAIL_REQUEST_FIELDS, []

    column_map = _email_column_map(rows)
    missing_fields = [
        field
        for field in EMAIL_REQUEST_FIELDS
        if field not in column_map and field not in EMAIL_COMPUTED_FIELDS
    ]
    if missing_fields and not allow_missing:
        raise MissingEmailFieldsError(missing_fields, _available_columns(rows))

    return EMAIL_REQUEST_FIELDS, [
        [_email_field_value(field, row, column_map) for field in EMAIL_REQUEST_FIELDS]
        for row in rows
    ]


def _email_field_value(field: str, row: dict[str, Any], column_map: dict[str, str]) -> Any:
    if field == "Bank Financed Y/N":
        return _bank_financed_value(row)
    if field == "Bot Error Comment":
        return _bot_error_comment(row)

    value: Any = ""
    if field == "Vehicle CV":
        value = row.get("VehicleValue") or row.get("Amount") or ""
    elif field == "Registration Emirates":
        value = row.get("Emirate") or row.get("RegistrationEmirates") or ""
    elif field == "Email":
        value = row.get("Email_CRM") or row.get("EmailCRM") or row.get("Email") or ""
    elif field == "Mobile Number":
        value = row.get("Mobile_CRM") or row.get("InsuredMobileNumber") or row.get("PhoneNumber") or ""
    elif field == "Bot Status(Passed/Error)":
        column = column_map.get(field)
        raw_value = row.get(column, "") if column else ""
        value = str(raw_value).upper() if raw_value is not None else ""
    else:
        column = column_map.get(field)
        if column:
            value = row.get(column, "")

    if field in EMAIL_DATE_FIELDS:
        return _format_date_only(value)

    return value


def _bot_error_comment(row: dict[str, Any]) -> Any:
    base_error_lines = _base_error_lines(row)
    missing_lines = _missing_reason_lines(row)
    if not base_error_lines and not missing_lines:
        return ""
    if CellRichText is None or TextBlock is None or InlineFont is None:
        return _plain_bot_error_comment(base_error_lines, missing_lines)
    return _rich_bot_error_comment(base_error_lines, missing_lines)


def _plain_bot_error_comment(base_error_lines: list[str], missing_lines: list[tuple[str, list[str]]]) -> str:
    lines: list[str] = []
    lines.extend(base_error_lines)
    for label, items in missing_lines:
        lines.append(f"{label}: {', '.join(items)}")
    return "\n".join(lines)


def _rich_bot_error_comment(base_error_lines: list[str], missing_lines: list[tuple[str, list[str]]]) -> CellRichText:
    rich_text = CellRichText()
    is_first = True
    for line in base_error_lines:
        prefix = "" if is_first else "\n"
        rich_text.append(TextBlock(_inline_font(ERROR_COMMENT_BASE_COLOR, bold=True), f"{prefix}{line}"))
        is_first = False
    for label, items in missing_lines:
        prefix = "" if is_first else "\n"
        rich_text.append(TextBlock(_inline_font(ERROR_COMMENT_LABEL_COLOR, bold=True), f"{prefix}{label}: "))
        rich_text.append(", ".join(items))
        is_first = False
    return rich_text


def _first_text(*values: Any) -> str:
    for value in values:
        text = str(value or "").strip()
        if text:
            return text
    return ""


def _base_error_lines(row: dict[str, Any]) -> list[str]:
    lines: list[str] = []
    for value in (row.get("ValidationError"), row.get("LastError")):
        text = str(value or "").strip()
        if not text:
            continue
        parts = [part.strip() for part in text.splitlines() if part.strip()]
        lines.extend(parts or [text])
    return lines


def _missing_reason_lines(row: dict[str, Any]) -> list[tuple[str, list[str]]]:
    final_status = str(row.get(LOG_FINAL_STATUS_KEY) or "").strip().casefold()
    if final_status == "prospect":
        return []
    if final_status == "lead":
        return _prospect_missing_lines(row, docs_label="Missing Documents For a Prospect")
    if not final_status:
        lines: list[tuple[str, list[str]]] = []
        lines.extend(_prospect_missing_lines(row, docs_label="Missing Documents For a Prospect"))
        lines.extend(_lead_missing_lines(row))
        return lines
    return []


def _prospect_missing_lines(row: dict[str, Any], docs_label: str) -> list[tuple[str, list[str]]]:
    lines: list[tuple[str, list[str]]] = []
    lines.extend(_missing_lines("Missing Fields For a Prospect", row.get(LOG_MISSING_PROSPECT_FIELDS_KEY)))
    lines.extend(_missing_lines(docs_label, row.get(LOG_MISSING_PROSPECT_DOCS_KEY)))
    return lines


def _lead_missing_lines(row: dict[str, Any]) -> list[tuple[str, list[str]]]:
    lines: list[tuple[str, list[str]]] = []
    lines.extend(_missing_lines("Missing Fields For a Lead", row.get(LOG_MISSING_LEAD_FIELDS_KEY)))
    lines.extend(_missing_lines("Missing Documents For a Lead", row.get(LOG_MISSING_LEAD_DOCS_KEY)))
    return lines


def _missing_lines(label: str, value: Any) -> list[tuple[str, list[str]]]:
    items = _split_missing_values(value)
    if not items:
        return []
    return [(label, items)]


def _split_missing_values(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    parts = [part.strip() for part in re.split(r"[,\n;]+", text) if part.strip()]
    return [_humanize_missing_item(part) for part in parts]


def _humanize_missing_item(value: str) -> str:
    text = value.strip()
    if not text:
        return text
    if text.replace(" ", "").isupper():
        return text
    text = text.replace("_", " ")
    text = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", text)
    text = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", text)
    return " ".join(text.split())


def _inline_font(color: str, bold: bool = False) -> InlineFont:
    if InlineFont is None:
        raise RuntimeError("InlineFont is not available in this openpyxl version.")
    return InlineFont(color=_argb_color(color), b=bold)


def _argb_color(value: str) -> str:
    cleaned = value.strip().lstrip("#").upper()
    if len(cleaned) == 6:
        return f"FF{cleaned}"
    return cleaned


def _plain_text(value: Any) -> str:
    if value is None:
        return ""
    plain = getattr(value, "plain", None)
    if isinstance(plain, str):
        return plain
    return str(value)


def _bank_financed_value(row: dict[str, Any]) -> str:
    bank_name = row.get("BankName")
    if isinstance(bank_name, str) and bank_name.strip():
        return "Y"

    payment_option = row.get("PaymentOption")
    if isinstance(payment_option, str):
        normalized = payment_option.strip().casefold()
        if normalized in {"bank", "financed", "finance", "loan", "yes", "y", "true", "1"}:
            return "Y"
        if normalized in {"cash", "no", "n", "false", "0"}:
            return "N"

    return ""


def _format_date_only(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text:
        return ""
    if len(text) >= 10 and text[4] == "-" and text[7] == "-":
        return text[:10]
    return text

def _email_column_map(rows: list[dict[str, Any]]) -> dict[str, str]:
    available_columns = _available_columns(rows)
    normalized_columns = {_normalize_column(column): column for column in available_columns}
    column_map: dict[str, str] = {}

    for field in EMAIL_REQUEST_FIELDS:
        candidates = [field, *EMAIL_FIELD_ALIASES.get(field, [])]
        for candidate in candidates:
            match = normalized_columns.get(_normalize_column(candidate))
            if match:
                column_map[field] = match
                break
    return column_map

def _available_columns(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for column in row:
            if column not in seen:
                seen.add(column)
                columns.append(column)
    return columns

def _normalize_column(value: str) -> str:
    return "".join(ch for ch in str(value).casefold() if ch.isalnum())

def _add_logo(ws, settings: Settings, cell: str) -> None:
    if ExcelImage is None:
        return
    logo_path = _resolve_logo_path(settings)
    if not logo_path.exists():
        return
    try:
        logo = ExcelImage(str(logo_path))
    except Exception:
        return

    target_height = 80
    aspect_ratio = logo.width / logo.height if logo.height else 1
    logo.height = target_height
    logo.width = int(target_height * aspect_ratio)
    logo.anchor = _image_anchor(cell, logo.width, logo.height, left_padding=18, top_padding=12)

    ws.add_image(logo)

def _image_anchor(cell: str, width: int, height: int, left_padding: int = 0, top_padding: int = 0) -> OneCellAnchor:
    coordinate = ws_coordinate(cell)
    marker = AnchorMarker(
        col=coordinate["col"],
        row=coordinate["row"],
        colOff=pixels_to_EMU(left_padding),
        rowOff=pixels_to_EMU(top_padding),
    )
    size = XDRPositiveSize2D(pixels_to_EMU(width), pixels_to_EMU(height))
    return OneCellAnchor(_from=marker, ext=size)

def ws_coordinate(cell: str) -> dict[str, int]:
    col_letters = "".join(ch for ch in cell if ch.isalpha())
    row_digits = "".join(ch for ch in cell if ch.isdigit())
    col_index = 0
    for char in col_letters.upper():
        col_index = col_index * 26 + (ord(char) - ord("A") + 1)
    return {"col": col_index - 1, "row": int(row_digits) - 1}

def _resolve_logo_path(settings: Settings) -> Path:
    raw_path = (settings.excel_logo_path or "").strip()
    normalized_path = raw_path.replace("\\", "/")
    path = Path(normalized_path)
    if path.is_absolute():
        return path
    return Path(__file__).resolve().parent.parent / path

def _style_cells(ws, cell_range: str, fill: str | None = None, border: str | None = None) -> None:
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = _fill(fill)
            if border:
                cell.border = _border(border)
            cell.alignment = Alignment(vertical="center")

def _fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)

def _border(color: str) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)

def _safe_sheet_name(value: str) -> str:
    return value[:31].replace("/", "-").replace("\\", "-").replace("?", "")
