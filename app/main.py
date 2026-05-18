from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import get_settings
from app.data_access import RequestFilters, STATUS_ORDER, get_repository, _normalize_status_counts

settings = get_settings()
repository = get_repository(settings)

BASE_DIR = Path(__file__).resolve().parent

app = FastAPI(title=settings.app_title)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))


def build_filters(
    request_id: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    status: Optional[str] = None,
    q: Optional[str] = None,
    page: int = 1,
    page_size: int = 25,
) -> RequestFilters:
    return RequestFilters(
        request_id=request_id or None,
        date_from=date_from,
        date_to=date_to,
        status=status or None,
        q=q or None,
        page=page,
        page_size=page_size,
    )


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request):
    return templates.TemplateResponse(
        "index.html",
        {
            "request": request,
            "app_title": settings.app_title,
            "data_source": settings.data_source,
        },
    )


@app.get("/api/requests")
def api_requests(
    request_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(25, ge=1, le=200),
):
    filters = build_filters(request_id, date_from, date_to, status, q, page, page_size)
    return repository.list_requests(filters)


@app.post("/api/requests/{request_id}/retry")
def retry_request(request_id: str):
    row = repository.retry_request(request_id)
    if not row:
        raise HTTPException(status_code=404, detail="Request ID not found")
    return {"message": "Request moved back to Pending", "row": row}


def _safe_sheet_name(value: str) -> str:
    return value[:31].replace("/", "-").replace("\\", "-").replace("?", "")


def _apply_table_style(ws, max_row: int, max_col: int) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="E5E7EB")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_length = 12
        for cell in ws[letter]:
            max_length = max(max_length, min(len(str(cell.value or "")), 45))
        ws.column_dimensions[letter].width = max_length + 2


def _build_excel(rows: list[dict], filters: RequestFilters) -> bytes:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    counts = {}
    for row in rows:
        status = str(row.get("CRMStatus") or "Unknown")
        counts[status] = counts.get(status, 0) + 1
    counts = _normalize_status_counts(counts)

    summary.append(["CRM Request Monitor Export"])
    summary.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary.append(["Total Requests", len(rows)])
    summary.append([])
    summary.append(["Applied Filters"])
    summary.append(["Request ID", filters.request_id or "All"])
    summary.append(["Date From", filters.date_from.isoformat() if filters.date_from else "All"])
    summary.append(["Date To", filters.date_to.isoformat() if filters.date_to else "All"])
    summary.append(["Status", filters.status or "All"])
    summary.append(["Search", filters.q or "All"])
    summary.append([])
    summary.append(["Status", "Count"])
    for status in STATUS_ORDER:
        summary.append([status, counts.get(status, 0)])
    for status, count in counts.items():
        if status not in STATUS_ORDER:
            summary.append([status, count])

    summary["A1"].font = Font(size=16, bold=True)
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top")
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 30

    data_sheet = wb.create_sheet(_safe_sheet_name("Requests"))
    columns = list(rows[0].keys()) if rows else ["No data matched the selected filters"]
    data_sheet.append(columns)
    for row in rows:
        data_sheet.append([row.get(column) for column in columns])

    _apply_table_style(data_sheet, max(1, data_sheet.max_row), max(1, data_sheet.max_column))

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


@app.get("/api/export")
def export_excel(
    request_id: Optional[str] = Query(None),
    date_from: Optional[date] = Query(None),
    date_to: Optional[date] = Query(None),
    status: Optional[str] = Query(None),
    q: Optional[str] = Query(None),
):
    filters = build_filters(request_id, date_from, date_to, status, q, page=1, page_size=200)
    rows = repository.export_rows(filters)
    content = _build_excel(rows, filters)
    filename = f"crm_requests_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    headers = {"Content-Disposition": f"attachment; filename={filename}"}
    return StreamingResponse(
        BytesIO(content),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
